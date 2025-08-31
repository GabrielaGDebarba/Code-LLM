import os
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Folder with JSON files
json_folder = "E:\\MESTRADO\\BIM A+\\00\\MODULO 7 - DISSERTATION\\Rhino_Grasshopper\\info_extraction\\prompt_2.4\\json\\"

# Column order (explicit to keep stable ordering in the final Excel)
COLUMNS_ORDER = [
    "Layout Option",
    "Requested Area",
    "Final Area",
    "Requested Rooms",
    "Output Rooms",
    "Requested Windows",
    "Output Windows",
    "Requested Noise Level",
    "Output Noise Level",
    "Requested Sun Hours (Winter)",
    "Output Sun Hours (Winter)",
    "Requested Sun Hours (Summer)",
    "Output Sun Hours (Summer)",
    "Excess Area?",
    "Windows Exported?",
    "Entrance Door Exported?",
    "Interior Doors Exported?",
    "Exterior Walls Exported?",
    "Interior Walls Exported?",
    "Door connectivily problem?",  
    "Status"
]

# Helper: try to extract an integer sort key from layout_option (robust)
def layout_sort_key(item):
    lo = item.get("layout_option")
    # If it's already a number, return it
    if isinstance(lo, (int, float)):
        try:
            return (int(lo), "")
        except Exception:
            pass
    # If it's a string containing a number, extract the first integer
    if lo is not None:
        s = str(lo)
        m = re.search(r'-?\d+', s)
        if m:
            try:
                return (int(m.group()), "")
            except Exception:
                pass
        # fallback to lexicographic (stable)
        return (float("inf"), s.lower())
    # If None, push to end
    return (float("inf"), "")

# Load all JSON files into memory first so we can sort by layout_option
layouts = []
for filename in os.listdir(json_folder):
    if not filename.lower().endswith(".json"):
        continue
    path = os.path.join(json_folder, filename)
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        # skip problematic files but continue processing others
        print(f"Warning: failed to read {filename}: {e}")
        continue
    # store the data along with filename for debugging if needed
    layouts.append({"file": filename, "data": data, "layout_option": data.get("layout_option")})

# Sort layouts by numeric layout_option when possible
layouts.sort(key=layout_sort_key)

# Build table_data in sorted order
table_data = []
for entry in layouts:
    data = entry["data"]

    layout_option = data.get("layout_option")
    requested_area = data.get("requested_area")
    final_area = data.get("final_area")

    # Arrays
    requested_rooms = data.get("requested_rooms", [])
    output_rooms = data.get("output_rooms", [])
    requested_windows = data.get("requested_windows", [])
    output_windows = data.get("output_windows", [])
    requested_noise_level = data.get("requested_noise_level", [])
    output_noise_level = data.get("output_noise_level", [])
    requested_sun_hours_w = data.get("requested_sun_hours_w", [])
    output_sun_hours_w = data.get("output_sun_hours_w", [])
    requested_sun_hours_s = data.get("requested_sun_hours_s", [])
    output_sun_hours_s = data.get("output_sun_hours_s", [])

    # Single fields
    has_excess_area = data.get("has_excess_area")
    windows_exported = data.get("windows_exported")
    entrance_door_exported = data.get("entrance_door_exported")
    interior_door_exported = data.get("interior_door_exported")
    exterior_walls_exported = data.get("exterior_walls_exported")
    interior_walls_exported = data.get("interior_walls_exported")
    door_connecting_problem = data.get("door_connecting_problem")  

    # Determine number of rows for this layout
    num_rows = max(
        len(requested_rooms), len(output_rooms),
        len(requested_windows), len(output_windows),
        len(requested_noise_level), len(output_noise_level),
        len(requested_sun_hours_w), len(output_sun_hours_w),
        len(requested_sun_hours_s), len(output_sun_hours_s),
        1  # ensure at least one row even if all arrays empty
    )

    for i in range(num_rows):
        if i == 0:
            row = {
                "Layout Option": layout_option,
                "Requested Area": requested_area,
                "Final Area": final_area,
                "Requested Rooms": requested_rooms[i] if i < len(requested_rooms) else "null",
                "Output Rooms": output_rooms[i] if i < len(output_rooms) else "null",
                "Requested Windows": requested_windows[i] if i < len(requested_windows) else "null",
                "Output Windows": output_windows[i] if i < len(output_windows) else "null",
                "Requested Noise Level": requested_noise_level[i] if i < len(requested_noise_level) else "null",
                "Output Noise Level": output_noise_level[i] if i < len(output_noise_level) else "null",
                "Requested Sun Hours (Winter)": requested_sun_hours_w[i] if i < len(requested_sun_hours_w) else "null",
                "Output Sun Hours (Winter)": output_sun_hours_w[i] if i < len(output_sun_hours_w) else "null",
                "Requested Sun Hours (Summer)": requested_sun_hours_s[i] if i < len(requested_sun_hours_s) else "null",
                "Output Sun Hours (Summer)": output_sun_hours_s[i] if i < len(output_sun_hours_s) else "null",
                "Excess Area?": has_excess_area,
                "Windows Exported?": windows_exported,
                "Entrance Door Exported?": entrance_door_exported,
                "Interior Doors Exported?": interior_door_exported,
                "Exterior Walls Exported?": exterior_walls_exported,
                "Interior Walls Exported?": interior_walls_exported,
                "Door connectivily problem?": door_connecting_problem,
                "Status": ""  # MUST remain blank
            }
        else:
            # For subsequent rows, single-value fields must stay blank (not "null")
            row = {
                "Layout Option": "",
                "Requested Area": "",
                "Final Area": "",
                "Requested Rooms": requested_rooms[i] if i < len(requested_rooms) else "",
                "Output Rooms": output_rooms[i] if i < len(output_rooms) else "",
                "Requested Windows": requested_windows[i] if i < len(requested_windows) else "",
                "Output Windows": output_windows[i] if i < len(output_windows) else "",
                "Requested Noise Level": requested_noise_level[i] if i < len(requested_noise_level) else "",
                "Output Noise Level": output_noise_level[i] if i < len(output_noise_level) else "",
                "Requested Sun Hours (Winter)": requested_sun_hours_w[i] if i < len(requested_sun_hours_w) else "",
                "Output Sun Hours (Winter)": output_sun_hours_w[i] if i < len(output_sun_hours_w) else "",
                "Requested Sun Hours (Summer)": requested_sun_hours_s[i] if i < len(requested_sun_hours_s) else "",
                "Output Sun Hours (Summer)": output_sun_hours_s[i] if i < len(output_sun_hours_s) else "",
                "Excess Area?": "",
                "Windows Exported?": "",
                "Entrance Door Exported?": "",
                "Interior Doors Exported?": "",
                "Exterior Walls Exported?": "",
                "Interior Walls Exported?": "",
                "Door connectivily problem?": "",
                "Status": ""  # remain blank
            }
        # Ensure order of keys matches COLUMNS_ORDER (for pandas)
        ordered_row = {col: row.get(col, "") for col in COLUMNS_ORDER}
        table_data.append(ordered_row)

    # Blank separator row (all empty strings) to separate layouts
    table_data.append({col: "" for col in COLUMNS_ORDER})

# Create DataFrame with explicit column ordering
df = pd.DataFrame(table_data, columns=COLUMNS_ORDER)

# Save to Excel
excel_file = "layout_table_prompt_2.4.xlsx"
df.to_excel(excel_file, index=False)

# --- Format with openpyxl ---
wb = load_workbook(excel_file)
ws = wb.active

# Header style
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define thin gray border
thin_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF")
)

# Detect separator rows BEFORE mutating body cells
separator_rows = set()
for r in range(2, ws.max_row + 1):
    # consider a row a separator only if all cells are None or empty string (after stripping)
    if all((c.value is None) or (isinstance(c.value, str) and c.value.strip() == "") for c in ws[r]):
        separator_rows.add(r)

# Headers that should NOT receive "null" fill (single-value columns + Status)
no_null_headers = {
    "Layout Option", "Requested Area", "Final Area",
    "Excess Area?", "Windows Exported?", "Entrance Door Exported?",
    "Interior Doors Exported?", "Exterior Walls Exported?", "Interior Walls Exported?",
    "Door connectivily problem?", "Status"
}

# Body style (only for data rows; skip separator rows)
for r in range(2, ws.max_row + 1):
    if r in separator_rows:
        continue
    for cell in ws[r]:
        header = ws.cell(row=1, column=cell.column).value
        # Normalize header type to str to compare safely
        header_str = str(header) if header is not None else ""
        # If this header is in the no_null list, never set cell to "null" (keep blank)
        if header_str in no_null_headers:
            # If there is an existing literal "null" string, replace it with blank
            if isinstance(cell.value, str) and cell.value.strip().lower() == "null":
                cell.value = ""
            # leave None or "" as blank
            if cell.value is None:
                cell.value = ""
            cell.font = Font(italic=False, color="000000")
        else:
            # Columns allowed to show "null" as placeholder when empty
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = "null"
                cell.font = Font(italic=True, color="808080")  # gray + italic
            elif isinstance(cell.value, str) and cell.value.strip().lower() == "null":
                # keep existing "null" but style it
                cell.value = "null"
                cell.font = Font(italic=True, color="808080")
            else:
                cell.font = Font(italic=False, color="000000")

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

# Auto-adjust column width
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# Color separator rows across all columns
fill_separator = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for r in sorted(separator_rows):
    for c in range(1, ws.max_column + 1):  
        cell = ws.cell(row=r, column=c)
        cell.value = ""  # keep empty
        cell.fill = fill_separator
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Ensure wrap_text is applied to all cells (header + body + separators)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical,
            wrap_text=True
        )

wb.save(excel_file)
print(f"File '{excel_file}' generated successfully with formatting applied!")
